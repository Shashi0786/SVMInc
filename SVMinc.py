import glob
import mimetypes
import os
import logging
import smtplib
import time
from datetime import datetime, timedelta, date, timezone
from email.message import EmailMessage
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from seleniumbase import BaseCase
from win32com.client import Dispatch


def vlookup(file_path_A, file_path_B, column_index_A, column_index_B):
    workbook_A = load_workbook(file_path_A)
    worksheet_A = workbook_A.active
    workbook_B = load_workbook(file_path_B)
    worksheet_B = workbook_B.active
    for row_A in range(2, worksheet_A.max_row + 1):
        lookup_value = worksheet_A.cell(row=row_A, column=column_index_A).value
        for row_B in range(2, worksheet_B.max_row + 1):
            cell_value = worksheet_B.cell(row=row_B, column=column_index_B).value
            if cell_value == lookup_value:
                matched_value = worksheet_B.cell(row=row_B, column=2).value
                worksheet_A.cell(row=row_A, column=6).value = matched_value
    workbook_B.close()
    workbook_A.save(file_path_A)


def Send_selenium_report(recipients, Body, directory, subject, filesName):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['To'] = ', '.join(recipients)
    msg['From'] = 'erpsupport@svmincorporation.com'
    msg.add_alternative('This is a PLAIN TEXT', subtype='plain')
    msg.add_alternative(Body, subtype='html')
    for file in filesName:
        path = os.path.join(directory, file)
        if not os.path.isfile(path):
            continue
        ctype, encoding = mimetypes.guess_type(path)
        if ctype is None or encoding is not None:
            ctype = 'application/octet-stream'
        maintype, subtype = ctype.split('/', 1)
        with open(path, 'rb') as fp:
            msg.add_attachment(fp.read(), maintype=maintype, subtype=subtype, filename=file)
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login('erpsupport@svmincorporation.com', 'Erpsupport#223344')
    server.send_message(msg)
    server.quit()


def Change_File_Ext(save_folder):
    excel = Dispatch("Excel.Application")
    extensions = (".xls",)
    for file in os.listdir(os.path.normpath(os.path.join(os.getcwd(), save_folder))):
        if file.endswith(extensions):
            filename, included_extensions = os.path.splitext(file)
            filename = filename.replace(':', '-')  # Replace ':' with '-'
            src_path = os.path.normpath(os.path.join(os.getcwd(), save_folder, file))
            dst_path = os.path.normpath(os.path.join(os.getcwd(), save_folder, filename + '.xlsx'))
            wb = excel.Workbooks.Open(src_path)
            wb.SaveAs(dst_path, FileFormat=51)
            wb.Close()
            excel.Quit()
            os.remove(src_path)


def rename_latest_file(source_folder, new_filename):
    # Validate source_folder
    if not os.path.isdir(source_folder):
        logging.error(f"'{source_folder}' is not a valid directory path.")
        return None

    source_folder = os.path.join(source_folder, '')
    files = glob.glob(source_folder + '*')

    if not files:
        logging.warning("No files found in the specified folder.")
        return None

    max_file = max(files, key=os.path.getmtime)
    if not os.path.isfile(max_file):
        logging.warning(f"{max_file} is not a regular file.")
        return None

    filename, file_extension = os.path.splitext(os.path.basename(max_file))
    new_directory = os.path.dirname(max_file)
    new_path = os.path.join(new_directory, new_filename + file_extension)

    try:
        os.rename(max_file, new_path)
        logging.info(f"File '{max_file}' renamed to '{new_filename + file_extension}'.")
        return new_path
    except Exception as e:
        logging.error(f"Failed to rename file: {e}")
        return None


def apply_Header_formatting(sheet, min_row):
    for X in range(1, sheet.max_column + 1):
        char = get_column_letter(X)
        cell = sheet[char + str(min_row)]
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.font = Font(bold=True, size='13', italic=False, color='00FFFFFF')
        cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             top=Side(border_style='double', color='FF000000'),
                             bottom=Side(border_style='double', color='FF000000'))


def apply_Sheet_formatting(sheet, min_row, min_col):
    for X in range(min_row, sheet.max_row + 1):
        for Y in range(min_col, sheet.max_column + 1):
            cell = sheet.cell(X, Y)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(bold=False, size=11, italic=True)
            cell.border = Border(left=Side(style='thin', color='FF000000'),
                                 right=Side(style='thin', color='FF000000'),
                                 top=Side(style='thin', color='FF000000'),
                                 bottom=Side(style='thin', color='FF000000'))


def apply_Alignment_formatting(sheet, min_row, min_col):
    for X in range(min_row, sheet.max_row + 1):
        for Y in range(min_col, sheet.max_column + 1):
            cell = sheet.cell(X, Y)
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def apply_Font_formatting(sheet, min_row, min_col):
    for X in range(min_row, sheet.max_row + 1):
        for Y in range(min_col, sheet.max_column + 1):
            cell = sheet.cell(X, Y)
            cell.font = Font(bold=False, size=11, italic=True)


def apply_Border_formatting(sheet, min_row, min_col):
    for X in range(min_row, sheet.max_row + 1):
        for Y in range(min_col, sheet.max_column + 1):
            cell = sheet.cell(X, Y)
            cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))


def apply_Value_formatting(sheet, min_row, min_col):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=min_col, max_col=sheet.max_column):
        for cell in row:
            cell.number_format = '_(* #,##0_);_(* #,##0_);_(* "-"??_);_(@_)'
            if cell.value is None or cell.value == '':
                cell.value = 0


def apply_CellBackground_formatting(sheet, min_row, Color):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type='solid')
    for cell in sheet[min_row]:
        cell.fill = fill


def apply_Date_Formating(sheet, min_row, max_col):
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=max_col, max_col=max_col):
        for cell in row:
            cell.number_format = 'dd-mmm-yyyy'
            if isinstance(cell.value, str):  # Check if the cell value is a string
                try:  # Format the cell value if it's in the expected format
                    parsed_date = datetime.strptime(cell.value, "%m/%d/%Y %I:%M:%S %p")
                    cell.value = parsed_date
                    cell.number_format = 'dd-mmm-yyyy'
                except ValueError:
                    pass  # Ignore if parsing fails


def apply_ChangeCase_Formatting(sheet, min_row, type):
    for row in sheet.iter_rows(min_row=min_row, max_row=min_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if type == 'U':
                cell.value = str(cell.value).upper()
            elif type == 'L':
                cell.value = str(cell.value).lower()
            elif type == 'P':
                cell.value = str(cell.value).title()


def apply_DeleteRow_BasedonValue_Comparison(sheet, min_row, min_col, max_col):
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row - 1):
        if isinstance(row[max_col].value, (int, float)) and isinstance(row[min_col].value, (int, float)):
            if row[max_col].value >= 0.9 * row[min_col].value:
                rows_to_delete.append(row[0].row)
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)


def removeFormatting(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'


def apply_DeleteRows_by_Condition(sheet: object, condition: object) -> object:
    rows_to_delete = [i for i in range(sheet.max_row, 1, -1) if condition(sheet, i)]
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)


def apply_fill_color(sheet, start_date, end_date):
    fillA = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
    fillB = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fillC = PatternFill(start_color="0000FF00", end_color="0000FF00", fill_type="solid")
    for i in range(2, sheet.max_row + 1):
        cell_date = datetime.strptime(sheet.cell(i, 3).value, "%Y-%m-%d") if isinstance(sheet.cell(i, 3).value,
                                                                                        str) else sheet.cell(i, 3).value
        if isinstance(cell_date, date):
            if cell_date < datetime.combine(start_date, datetime.min.time()):
                sheet.cell(row=i, column=3).fill = fillA
            elif datetime.combine(start_date, datetime.min.time()) <= cell_date <= datetime.combine(end_date,
                                                                                                    datetime.min.time()):
                sheet.cell(row=i, column=3).fill = fillB
            elif cell_date > datetime.combine(end_date, datetime.min.time()):
                sheet.cell(row=i, column=3).fill = fillC
            else:
                continue


def apply_SortSheet_by_date(destination_sheet, SortingHeader):
    data = list(destination_sheet.iter_rows(values_only=True))
    cols = data[0]
    df = pd.DataFrame(data[1:], columns=cols)
    df[SortingHeader] = pd.to_datetime(df[SortingHeader])
    df.sort_values(by=SortingHeader, inplace=True)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            destination_sheet.cell(row=r_idx, column=c_idx, value=value)


def apply_AutoFit_Column(sheet):
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 1) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width


def apply_deleteColumns_ByHeaders(sheet, header_list):
    column_to_delete = []
    for i in range(sheet.max_column, 0, -1):
        if sheet.cell(1, i).value in header_list:
            column_to_delete.append(i)
    for col_index in sorted(column_to_delete, reverse=True):
        sheet.delete_cols(col_index)


def apply_ClearUnwanted_Text(sheet, minr, minc, maxc):
    for row in sheet.iter_rows(min_row=minr, max_row=sheet.max_row, min_col=minc, max_col=maxc):
        for cell in row:
            if cell.value == 'NOT ASSIGNED YET' or cell.value == 'NOT ASSIGNE':
                cell.value = None


def copy_specific_columns(source_ws, destination_ws, columns_to_copy):
    try:
        header_row = next(
            source_ws.iter_rows(min_row=1, max_row=1, values_only=True))  #print("Header Row:", header_row)
        for index, column_name in enumerate(columns_to_copy, start=1):  #print("Column Name to Copy:", column_name)
            column_index = header_row.index(column_name) + 1
            destination_ws.cell(row=1, column=index).value = column_name
            for row_index, cell in enumerate(source_ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index),
                                             start=2):
                destination_ws.cell(row=row_index, column=index).value = cell[0].value
    except Exception as e:
        print("Error:", e)


def generate_sumifs_formula(source_wb, detail_sheet_name, SumeRange_Col, CriteriaRange_Col1, summary_sheet_name,
                            summary_column, CriteriaRange_Col2=None, Criteria2=None):
    Row = source_wb[summary_sheet_name].max_row
    last_row = source_wb[detail_sheet_name].max_row
    detail_range = f"'{detail_sheet_name}'!${SumeRange_Col}$1:${SumeRange_Col}${last_row}"
    criteria_range1 = f"'{detail_sheet_name}'!${CriteriaRange_Col1}$1:${CriteriaRange_Col1}${last_row}"
    criteria1 = f"'{summary_sheet_name}'!{summary_column}2:{summary_column}{Row}"  # Assuming A2 is the starting row for criteria
    formula = f"=SUMIFS({detail_range}, {criteria_range1}, {criteria1}"
    if CriteriaRange_Col2 is not None and Criteria2 is not None:
        criteria_range2 = f"'{detail_sheet_name}'!${CriteriaRange_Col2}$1:${CriteriaRange_Col2}${last_row}"
        criteria2 = f"'{summary_sheet_name}'!{Criteria2}"
        formula += f", {criteria_range2}, {criteria2}"
    formula += ")"
    return formula


def apply_Remove_duplicates(ws, StartRow):
    max_row = ws.max_row
    max_col = ws.max_column
    unique_rows = set()
    for row in range(max_row, StartRow, -1):
        row_values = [ws.cell(row=row, column=col).value for col in range(StartRow, max_col + 1)]
        row_tuple = tuple(row_values)
        if row_tuple not in unique_rows:
            unique_rows.add(row_tuple)
        else:
            ws.delete_rows(row)


def Procurement_Status_Report_Formating(SOURCE_FILE_PATH):
    wb = load_workbook(SOURCE_FILE_PATH)
    ws = wb.active

    # Delete specific columns by headers
    headers_to_delete = ["SELECT", "ITEM_GROUP_NAME", 'ORDER_DATE', 'SHIPPED_STATUS', "Recv_Percentage", 'ORDER_EX_QTY',
                         "BAL", 'STOCK_TRANSFER_LOCATION', 'Balance_To_Transfer', 'RECV_FROM_JOBWORK', "RAW_ITEM_NAME",
                         "reprocess_issue_qty", "reprocess_recv_qty", 'MERCHANT_ID', 'SEASON', 'ORDER_EX_QTY',
                         'TOTAL_AVERAGE', 'RECV_FROM_JOBWORK', "reprocess_bal_qty", "Party_Payment_Term",
                         "balance_to_issue", 'UOM', 'PO_STATUS', 'Extend_Date', 'BALANCE_IN_HAND_VALUE',
                         'last_ship_date', 'BOM_AMOUNT', 'PO_AMOUNT', 'AMOUNT_DIFF']
    apply_deleteColumns_ByHeaders(ws, headers_to_delete)

    # Remove duplicates
    apply_Remove_duplicates(ws, 1)

    # Clear unwanted text
    apply_ClearUnwanted_Text(ws, 1, 8, 20)

    # Apply sheet formatting
    apply_Sheet_formatting(ws, 2, 1)

    # Apply header formatting
    apply_Header_formatting(ws, 1)

    # Apply cell background formatting
    apply_CellBackground_formatting(ws, 1, '000000FF')

    # Autofit columns
    apply_AutoFit_Column(ws)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['T'].width = 18

    # Set autofilter
    FullRange = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.auto_filter.ref = FullRange

    # Save and close workbook
    wb.save(SOURCE_FILE_PATH)
    wb.close()

    # Pause for 1 second
    time.sleep(1)


def Order_Completion_Report_Formating(SOURCE_FILE_PATH, DESTINATION_FILE_PATH):
    source_wb = load_workbook(SOURCE_FILE_PATH)
    source_ws = source_wb.active
    destination_workbook = Workbook()
    destination_sheet = destination_workbook.active

    # Copy specific columns from source to destination
    columns_to_copy = ['Order No', 'Order Date', 'Delivery Date', 'Buyer', 'Style Name',
                       'Order Qty', 'Cost Price', 'Cut Qty', 'Stitch Qty', 'Finished Qty',
                       'Packed Qty', 'Good Pcs To Pack', 'Invoice No', 'Invoice Qty']
    copy_specific_columns(source_ws, destination_sheet, columns_to_copy)

    # Remove duplicates
    apply_Remove_duplicates(destination_sheet, 1)

    # Define start and end date for date filtering
    start_date = datetime.now(timezone.utc).date()
    end_date = start_date + timedelta(days=12)

    # Delete rows based on condition
    apply_DeleteRows_by_Condition(destination_sheet, lambda sheet, i: source_ws.cell(i, 4).value in (
        'NUEVOSDAMAS', "NUEVOSPORTA", "ARIELLE SOURCING PVT LTD", "URBAN OUTFITTERS", "ANTHROPOLOGIE",
        "DISRIBUIDORA LIVEPOOL, S.A. DE.C.V", "JY HOME ( ZHEJIANG )COM.LTD HARBOR HOUSE", "KAS AUSTRALIA PTY LTD",
        "DEVI DESIGNS LLC", "DORMIFY KEECO LLC"))

    # Apply header and cell background formatting
    apply_Header_formatting(destination_sheet, 1)
    apply_CellBackground_formatting(destination_sheet, 1, '000000FF')

    # Apply sheet formatting
    apply_Sheet_formatting(destination_sheet, 2, 1)

    # Apply value and date formatting
    apply_Value_formatting(destination_sheet, 2, 7)
    apply_Date_Formating(destination_sheet, 2, 2)
    apply_Date_Formating(destination_sheet, 2, 3)

    # Sort sheet by delivery date
    apply_SortSheet_by_date(destination_sheet, 'Delivery Date')

    # Apply fill color
    apply_fill_color(destination_sheet, start_date, end_date)

    # Delete rows based on value comparison
    apply_DeleteRow_BasedonValue_Comparison(destination_sheet, 2, 5, 13)

    # Autofit columns
    apply_AutoFit_Column(destination_sheet)

    # Set autofilter
    FullRange = f"A1:{get_column_letter(destination_sheet.max_column)}{destination_sheet.max_row}"
    destination_sheet.auto_filter.ref = FullRange

    # Save and close destination workbook
    destination_workbook.save(DESTINATION_FILE_PATH)
    source_wb.close()

    # Pause for 1 second
    time.sleep(1)

    # Remove source file
    os.remove(SOURCE_FILE_PATH)


def PO_Followup_Report_Formating(SOURCE_FILE_PATH, DESTINATION_FILE_PATH):
    wb = load_workbook(SOURCE_FILE_PATH)
    sheet = wb.active
    destination_workbook = Workbook()
    ws = destination_workbook.active

    # Copy data from source to destination workbook
    for row in sheet.iter_rows():
        for cell in row:
            ws[cell.coordinate].value = cell.value

    # Delete specific columns by headers
    headers_to_delete = ['s no', 'po type', 'approval status', 'delivery date', 'rp iss qty',
                         'rp recd qty', 'rp bal qty', 'rp return qty', 'final bal with rp qty with vendor',
                         'po status', 'show po', 'debit qty', 'debit note no', 'to be rcv giq rawconsumed',
                         'net greige issued', 'raw bal iss qty value', 'raw debit bal qty', 'n next days',
                         'raw actual consume qty', 'raw debit bal value']
    apply_deleteColumns_ByHeaders(ws, headers_to_delete)

    # Remove duplicates
    apply_Remove_duplicates(ws, 1)

    # Apply header formatting
    apply_Header_formatting(ws, 1)

    # Apply cell background formatting
    apply_CellBackground_formatting(ws, 1, '000000FF')

    # Delete rows based on value comparison
    apply_DeleteRow_BasedonValue_Comparison(ws, 2, 11, 12)

    # Apply sheet formatting
    apply_Sheet_formatting(ws, 2, 1)

    # Apply value formatting
    apply_Value_formatting(ws, 2, 8)

    # Apply date formatting
    apply_Date_Formating(ws, 2, 2)
    apply_Date_Formating(ws, 2, 7)

    # Apply change case formatting
    apply_ChangeCase_Formatting(ws, 1, 'U')

    # Autofit columns
    apply_AutoFit_Column(ws)
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['P'].width = 40
    ws.column_dimensions['Z'].width = 30

    # Set autofilter
    FullRange = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.auto_filter.ref = FullRange

    # Save destination workbook
    destination_workbook.save(DESTINATION_FILE_PATH)
    wb.close()

    # Pause for 1 second
    time.sleep(1)

    # Remove source file
    os.remove(SOURCE_FILE_PATH)


def Production_Detail_Report_Formating(inputFile_Path):
    wb = load_workbook(inputFile_Path)
    ws = wb.active

    # Unmerge cells and delete columns
    ws.unmerge_cells("A1:J1")
    ws.delete_cols(21, 6)

    # Move and delete rows
    ws.move_range("A4:E5", rows=-1)
    ws.move_range("A3:E5", cols=5)
    ws.delete_rows(ws.max_row)

    # Remove duplicates
    apply_Remove_duplicates(ws, 7)

    # Calculate SUBTOTAL in each column
    rows = 6
    for X in range(4, ws.max_column):
        char = get_column_letter(X)
        ws[char + str(rows)] = f"=SUBTOTAL(9,{char + '8'}:{char + str(ws.max_row)})"

    # Set autofilter
    FullRange = "A7:" + get_column_letter(ws.max_column) + str(ws.max_row)
    ws.auto_filter.ref = FullRange

    # Autofit columns
    apply_AutoFit_Column(ws)

    # Apply sheet formatting
    apply_Sheet_formatting(ws, 6, 1)

    # Delete rows based on value comparison
    apply_DeleteRow_BasedonValue_Comparison(ws, 8, 3, 17)

    # Apply header formatting
    apply_Header_formatting(ws, 7)
    apply_Header_formatting(ws, 6)

    # Apply cell background formatting
    apply_CellBackground_formatting(ws, 7, '000000FF')
    apply_CellBackground_formatting(ws, 6, '00666699')
    apply_CellBackground_formatting(ws, 1, '00000000')

    # Apply value formatting
    apply_Value_formatting(ws, 6, 3)

    # Set total label and merge cells
    ws.cell(6, 1).value = "TOTAL"
    ws.merge_cells("A6:C6")
    ws.merge_cells("A1:T1")

    # Save and close workbook
    wb.save(inputFile_Path)
    wb.close()

    # Pause for 1 second
    time.sleep(1)


def FormatingStockLedger(inputFile_Path, outputFile_path):
    # Rename File
    os.rename(inputFile_Path, outputFile_path)

    # Assigning Output File to Workbook
    wb = load_workbook(outputFile_path)
    ws = wb.active

    # Move and delete columns
    ws.move_range("A1:D6", cols=7)
    ws.delete_cols(1, 3)

    # Unhide specific columns
    ws.column_dimensions['E'].hidden = False
    ws.column_dimensions['J'].hidden = False

    # Apply sheet formatting
    apply_Sheet_formatting(ws, 7, 1)

    # Apply value formatting
    apply_Value_formatting(ws, 7, 3)

    # Clear cell value
    ws.cell(2, 2).value = ""

    # Calculate SUBTOTAL in each column
    rows = 7
    row = len(ws['A']) - 1
    for X in range(3, ws.max_column + 1):
        char = get_column_letter(X)
        ws[char + str(rows)] = f"=SUBTOTAL(9,{char + '9'}:{char + str(row)})"

    # Set total label
    ws.cell(7, 1).value = "TOTAL"
    ws.merge_cells("A7:B7")

    # Apply header formatting
    apply_Header_formatting(ws, 7)
    apply_Header_formatting(ws, 8)

    # Apply cell background formatting
    apply_CellBackground_formatting(ws, 7, '00666699')
    apply_CellBackground_formatting(ws, 8, '000000FF')

    # Set font formatting
    ws['H1'].font = Font(bold=False)

    # Set autofilter
    FullRange = "A8:" + get_column_letter(ws.max_column) + str(ws.max_row)
    ws.auto_filter.ref = FullRange

    # Autofit columns
    apply_AutoFit_Column(ws)
    ws.column_dimensions['A'].width = 45

    # Save and close workbook
    wb.save(outputFile_path)
    wb.close()

    # Pause for 1 second
    time.sleep(1)


def FormatingOrderInHand(File_Path):
    file_path_A = "downloaded_files\\sourceWB.xlsx"
    file_path_B = "downloaded_files\\ReportData.xlsx"
    column_index_A = 3
    column_index_B = 1

    # Call vlookup function
    vlookup(file_path_A, file_path_B, column_index_A, column_index_B)

    # Load destination workbook
    destination_workbook = load_workbook(file_path_A)
    destination_sheet = destination_workbook.active

    # Create source workbook
    source_wb = Workbook()
    DetailSheet = source_wb.active
    DetailSheet.title = 'InHand Detail'

    # Copy specific columns
    columns_to_copy = ['BUYER', 'STYLE', 'ORDER NO', 'COLOR', 'merchant', 'BUYER ORDER NO', 'DELIVERY DATE',
                       'CONV RATE', 'ORDER QTY', 'EXCESS PER', 'EXCESS QUANTITY', 'AVERAGE PRICE', 'AMOUNT FC',
                       'AMOUNT RS', 'ship qty', 'balance ship', 'buyer po no']
    copy_specific_columns(destination_sheet, DetailSheet, columns_to_copy)

    # Delete specified values
    values_to_delete = ('NUEVOSDAMAS', 'NUEVOSPORTA', 'NUEVOSKIDDOS')
    rows_to_delete = list(DetailSheet.iter_rows())
    for row in reversed(rows_to_delete):
        if row[0].value in values_to_delete:
            DetailSheet.delete_rows(row[0].row)

    # Apply date formatting
    apply_Date_Formating(DetailSheet, 2, 7)

    # Apply sheet formatting
    apply_Sheet_formatting(DetailSheet, 2, 1)

    # Autofit columns
    apply_AutoFit_Column(DetailSheet)

    # Apply header formatting
    apply_Header_formatting(DetailSheet, 1)

    # Apply cell background formatting
    apply_CellBackground_formatting(DetailSheet, 1, '000000FF')

    # Create SummarySheet
    SummarySheet = source_wb.create_sheet('InHand Summary')

    # Copy specific columns to SummarySheet
    columns_to_copy = ['BUYER']
    copy_specific_columns(DetailSheet, SummarySheet, columns_to_copy)

    # Remove duplicates in SummarySheet
    apply_Remove_duplicates(SummarySheet, 1)

    # Set cell value in SummarySheet
    SummarySheet["B1"].value = "Quantity"

    # Insert rows in DetailSheet
    DetailSheet.insert_rows(1)

    # Calculate SUBTOTAL in DetailSheet
    rows = 1
    for X in range(9, DetailSheet.max_column):
        char = get_column_letter(X)
        DetailSheet[char + str(rows)] = f"=SUBTOTAL(9,{char + '3'}:{char + str(DetailSheet.max_row)})"

    # Set autofilter in DetailSheet
    FullRange = "A2:" + get_column_letter(DetailSheet.max_column) + str(DetailSheet.max_row)
    DetailSheet.auto_filter.ref = FullRange

    # Apply header formatting and cell background formatting in DetailSheet
    apply_Header_formatting(DetailSheet, 1)
    apply_CellBackground_formatting(DetailSheet, 1, '00666699')

    # Apply value formatting in DetailSheet
    apply_Value_formatting(DetailSheet, 1, 8)

    # Extract data for transposing
    data = []
    for row in DetailSheet.iter_rows(min_row=3, max_row=DetailSheet.max_row, min_col=17, max_col=17, values_only=True):
        data.extend(row)

    # Convert data to DataFrame and remove duplicates
    df = pd.DataFrame(data, columns=['Values'])
    df.drop_duplicates(inplace=True)

    # Transpose data
    transposed_data = df.T.values.tolist()

    # Paste transposed data into SummarySheet
    for r_idx, row in enumerate(transposed_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            SummarySheet.cell(row=r_idx, column=c_idx + 2, value=value)

    # Insert rows in SummarySheet
    SummarySheet.insert_rows(1)

    # Generate and set formulas for each column in SummarySheet
    column_headers = ['C$2', 'D$2', 'E$2', 'F$2']
    detail_columns = ['Q', 'Q', 'Q', 'Q']
    result_columns = ['C', 'D', 'E', 'F']
    for col_idx, criteria2 in enumerate(column_headers):
        formula = generate_sumifs_formula(source_wb, 'InHand Detail', 'M', 'A', 'InHand Summary', 'A',
                                          detail_columns[col_idx], criteria2)
        for row in range(3, SummarySheet.max_row + 1):
            SummarySheet[f"{result_columns[col_idx]}{row}"].value = formula

    # Set formula for column B in SummarySheet
    formula = generate_sumifs_formula(source_wb, 'InHand Detail', 'P', 'A', 'InHand Summary', 'A', None, None)
    for row in range(3, SummarySheet.max_row + 1):
        SummarySheet[f"B{row}"].value = formula

    # Set formula for column G in SummarySheet
    formula = generate_sumifs_formula(source_wb, 'InHand Detail', 'N', 'A', 'InHand Summary', 'A', None, None)
    for row in range(3, SummarySheet.max_row + 1):
        SummarySheet[f"G{row}"].value = formula

    # Set autofilter in SummarySheet
    FullRange = "A2:" + get_column_letter(SummarySheet.max_column) + str(SummarySheet.max_row)
    SummarySheet.auto_filter.ref = FullRange

    # Apply header and cell background formatting in SummarySheet
    apply_Header_formatting(SummarySheet, 1)
    apply_CellBackground_formatting(SummarySheet, 1, '00666699')

    # Apply value formatting in SummarySheet
    apply_Value_formatting(SummarySheet, 1, 2)

    # Calculate SUBTOTAL in SummarySheet
    rows = 1
    for X in range(2, SummarySheet.max_column + 1):
        char = get_column_letter(X)
        SummarySheet[char + str(rows)] = f"=SUBTOTAL(9,{char + '3'}:{char + str(SummarySheet.max_row)})"

    # Set cell value in SummarySheet
    SummarySheet["G2"].value = "Total Value (INR)"

    # Apply sheet formatting in SummarySheet
    apply_Sheet_formatting(SummarySheet, 2, 1)

    # Set fixed width for columns in SummarySheet
    for column in SummarySheet.columns:
        SummarySheet.column_dimensions[column[0].column_letter].width = 20

    # Apply header formatting and cell background formatting in SummarySheet
    apply_Header_formatting(SummarySheet, 2)
    apply_CellBackground_formatting(SummarySheet, 2, '000000FF')

    # Save the workbook
    source_wb.active = source_wb['InHand Summary']
    source_wb.save(File_Path + '\\Order InHand Report.xlsx')


class NomeTest(BaseCase):
    def login_to_swag_labs(self):
        self.maximize_window()
        self.open('https://companyname.bluekaktus.com/#/login/:company')
        self.send_keys('#username', username)
        self.send_keys('#password', password)
        self.click('.submitForm')
        if self.assert_element_not_present("#company"):
            self.click_xpath("//span[text()='Confirm']")
            pass
        time.sleep(1)
        self.click('#company')
        self.click('#company > div > ul > li:nth-child(2) > a')
        self.click('#location')
        self.click('#location > div > ul > li:nth-child(2) > a')
        self.click('.submitForm')

    def select_target_date_Jquery(self, Element):
        if datetime.today().day == 1:
            first_day_of_current_month = datetime.today().replace(day=1)
            last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
            target_date = last_day_of_previous_month
            self.click('button[type="button"].previous > span')
        elif datetime.today().strftime("%#A") == 'Monday':
            target_date = datetime.today() - timedelta(days=2)
        else:
            target_date = datetime.today() - timedelta(days=1)
        target_day = target_date.strftime("%#d")
        for elm in self.find_elements(Element):
            if elm.text == target_day:
                elm.click()
                break

    def select_target_date_Basic(self, Element):
        if datetime.today().day == 1:
            first_day_of_current_month = datetime.today().replace(day=1)
            last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
            target_date = last_day_of_previous_month
            self.click("div#containerRow__2 th:nth-child(1) > button[type=\'button']")
        elif datetime.today().strftime("%A") == 'Monday':
            target_date = datetime.today() - timedelta(days=2)
        else:
            target_date = datetime.today() - timedelta(days=1)
        target_day = target_date.strftime("%d")
        for elm in self.find_elements(Element):
            if elm.text == target_day:
                elm.click()
                break

    manMenu = {"Accounts 1": "Set Up",
               "Order 2": "Set Up",
               "Master 1": "Create",
               "Order 1": "Create",
               "Accounts": "Transaction",
               "Procurement": "Transaction",
               "Production": "Transaction",
               "Master": "Transaction",
               "Reports": "Transaction",
               "Admin": "Transaction",
               "Order": "Transaction",
               "Permission": "Transaction",
               "Commercial": "Transaction",
               "API": "Transaction",
               "TNA": "Transaction",
               "Product Development": "Transaction",
               "Report Tool View": "Transaction"}

    subMenu = {"Chart of Accounts": "Accounts 1",
               "Chart of Cost Center": "Accounts 1",
               "Order Attribute": "Order 2",
               "Attribute Group Master": "Master 1",
               "Detail Item Master": "Master 1",
               "Summary Item Master": "Master 1",
               "Product Category Group": "Order 1",
               "Product Category": "Order 1",
               "Sales Order Group": "Order 1",
               "Account Configure": "Accounts",
               "Advance Payable": "Accounts",
               "Advance Receipt": "Accounts",
               "Authentication": "Accounts",
               "Balance Sheet": "Accounts",
               "Bank Master": "Accounts",
               "Bank Reconciliation": "Accounts",
               "Bill Paid Status": "Accounts",
               "Bill Posting": "Accounts",
               "Buyer Ageing": "Accounts",
               "Buyer Bill Status": "Accounts",
               "Buyer Ledger Currency Wise": "Accounts",
               "Cash Purchase": "Accounts",
               "Cost Center Ledger": "Accounts",
               "Cost Center Ledgers": "Accounts",
               "Cost Center Report": "Accounts",
               "Credit Note": "Accounts",
               "Day Book Register": "Accounts",
               "Debit Note": "Accounts",
               "Debit/Credit Register Report": "Accounts",
               "Export Bills": "Accounts",
               "GST Register": "Accounts",
               "GST Summary Register": "Accounts",
               "HSN Sale Account Mapping": "Accounts",
               "HSN Wise Summary": "Accounts",
               "IGST Payment Receipt": "Accounts",
               "Import Loan Opening Balance": "Accounts",
               "Incentive Account Port Wise Mapping": "Accounts",
               "Ledger (New)": "Accounts",
               "Loan / Pcl Entry": "Accounts",
               "Map Activity With Expense": "Accounts",
               "Multiple Bill Printing": "Accounts",
               "Multiple Voucher Printing": "Accounts",
               "Opening Balance": "Accounts",
               "Payment Audit Verification": "Accounts",
               "Payment HOD Verification": "Accounts",
               "Payment Release": "Accounts",
               "Pending Bill Report": "Accounts",
               "Petty Cash Posting": "Accounts",
               "Petty Cash UnPosting": "Accounts",
               "Petty Cash Voucher": "Accounts",
               "Production Bill": "Accounts",
               "Profit and Loss Account": "Accounts",
               "Purchase Register (Item Wise)": "Accounts",
               "Sale Register (Item Wise)": "Accounts",
               "Sale Register Report": "Accounts",
               "Sale Summary Register": "Accounts",
               "Service PO Bill": "Accounts",
               "Stock Updation": "Accounts",
               "Stock Value": "Accounts",
               "TDS CONFIGURATION": "Accounts",
               "TDS Report": "Accounts",
               "Trans Type Mapping": "Accounts",
               "Trial Balance": "Accounts",
               "Trial Balance (Section Wise)": "Accounts",
               "Trial Balance New": "Accounts",
               "User Check List": "Accounts",
               "Vendor Bill Status": "Accounts",
               "Vendors Ageing": "Accounts",
               "Purchase Bill": "Accounts",
               "Jobwork Bill": "Accounts",
               "Sale Bill": "Accounts",
               "Journal Entry": "Accounts",
               "Bill Payment": "Accounts",
               "Payment Receipt": "Accounts",
               "Accessories/ Fabric Stock Report": "Procurement",
               "Auto Bom Planning": "Procurement",
               "BOM Planning Approval / Indent Generation": "Procurement",
               "Cutting Entry": "Procurement",
               "Daily allocation reports": "Procurement",
               "Dye Lot Master": "Procurement",
               "Extra Material Request Indent": "Procurement",
               "Gate Entry Report": "Procurement",
               "Gate Out Entry": "Procurement",
               "General Issue": "Procurement",
               "General Issue Approval": "Procurement",
               "General Receive": "Procurement",
               "GST Mapping": "Procurement",
               "GST Template": "Procurement",
               "Harmonized Details": "Procurement",
               "Indent Approval Item Wise": "Procurement",
               "Indent Status": "Procurement",
               "Independent Inward Gate entry": "Procurement",
               "Independent Outward Gate entry": "Procurement",
               "Inspection Confirmation": "Procurement",
               "inspection update": "Procurement",
               "Internal Costing Approval": "Procurement",
               "Internal Stock Transfer": "Procurement",
               "Issue Receipt of fabric": "Procurement",
               "Item With High Rate Variance": "Procurement",
               "Manufacturing PO": "Procurement",
               "Material Issue for Production": "Procurement",
               "Material Payment Schedule": "Procurement",
               "Material Return Production": "Procurement",
               "Minimum Stock Level Item Report": "Procurement",
               "Negative Stock": "Procurement",
               "On Time Delivery": "Procurement",
               "Order Item Wise PO Details": "Procurement",
               "Order ItemWise Po Details": "Procurement",
               "Pending Gate Entries": "Procurement",
               "Pending Inspection Report": "Procurement",
               "Pending Receiving for Billing": "Procurement",
               "PO Followup Report": "Procurement",
               "PO Treatment For Excess Receiving": "Procurement",
               "Procurement Dashboard": "Procurement",
               "Procurement Status": "Procurement",
               "Procurement Status New": "Procurement",
               "Purchase Order Status": "Procurement",
               "PURCHASE ORDER STATUS WEEKLY": "Procurement",
               "Purchase Order Transaction": "Procurement",
               "Quotation Comparison": "Procurement",
               "Quotation Comparison StyleOrder Wise": "Procurement",
               "Raw Material Ageing Report": "Procurement",
               "Raw Material Pending Po": "Procurement",
               "Re Process Issue": "Procurement",
               "Receiving Details Report": "Procurement",
               "return process to vendor": "Procurement",
               "Service Po": "Procurement",
               "Service PO Approval": "Procurement",
               "Stock Adjustment": "Procurement",
               "Stock Ledger (New)": "Procurement",
               "Style Wise Purchase Summary Report": "Procurement",
               "Transaction with Initial Items": "Procurement",
               "Vendor Payment": "Procurement",
               "Yarn Order Status Report": "Procurement",
               "Indent": "Procurement",
               "Request for Quotation": "Procurement",
               "Quotation": "Procurement",
               "Purchase Order": "Procurement",
               "Jobwork Order": "Procurement",
               "Purchase Order Approval": "Procurement",
               "Receiving": "Procurement",
               "Inspection": "Procurement",
               "Return": "Procurement",
               "Material Issue": "Procurement",
               "Material Return": "Procurement",
               "Stock Transfer Issue": "Procurement",
               "Stock Transfer Receive": "Procurement",
               "Accessories Daily Issue Recieve Register": "Production",
               "Accessory Issue Report Daily": "Production",
               "Actual Lot Cut Daily": "Production",
               "Auto Lot Planning": "Production",
               "Book Consumption": "Production",
               "Bulk Production Purchase Order": "Production",
               "Costing Report": "Production",
               "Daily Production Report": "Production",
               "Daily Production Summary Report": "Production",
               "Excess Fabric Issue Report": "Production",
               "Extra Fabric Issue Report": "Production",
               "Fabric Issue Against Lotcut Report": "Production",
               "Issue / consume Details": "Production",
               "Lot Cut Planning": "Production",
               "Lot Planned Report": "Production",
               "Lot Wise Cutting Report": "Production",
               "Process Map": "Production",
               "Prod Transaction WIP Issue": "Production",
               "Prod Transaction WIP Receive": "Production",
               "Production Consolidated Report": "Production",
               "Production Detail Report": "Production",
               "Production Detail Report Size & color Wise": "Production",
               "Production Detail Report Size Wise": "Production",
               "Production Inspection": "Production",
               "Production Inspection Report": "Production",
               "Production Pending Bills": "Production",
               "Production Planning": "Production",
               "Production Po Report": "Production",
               "Production Po Status": "Production",
               "Production Process Planning": "Production",
               "Production Purchase Order": "Production",
               "Production Purchase Order Approval": "Production",
               "Production Receive": "Production",
               "Production Status Report": "Production",
               "Production Summary Report": "Production",
               "Production Transactions": "Production",
               "Stock Valuation Report": "Production",
               "WIP Stock Report": "Production",
               "WIP Transfer Query": "Production",
               "ACCESSORY DETAIL MASTER": "Master",
               "Analytics": "Master",
               "Auto Ledger Configuration": "Master",
               "Barcode Printing": "Master",
               "Block Library": "Master",
               "Brand Master": "Master",
               "Bulk Stock Allocation": "Master",
               "Carton Master": "Master",
               "Currency Master": "Master",
               "Currency Rate Master": "Master",
               "Delay Reason Master": "Master",
               "Delivery Terms Account Mapping": "Master",
               "Design Group": "Master",
               "Design Master": "Master",
               "Export Invoice Hsn Wise": "Master",
               "Export Sale Data Invoice": "Master",
               "Finencial Dash Borad": "Master",
               "Fixed List Updation": "Master",
               "Forwarder master": "Master",
               "GENERAL ITEM MASTER": "Master",
               "General Type Master": "Master",
               "GST Return Detail": "Master",
               "Item Catalog": "Master",
               "Item Excess": "Master",
               "Item Group Type": "Master",
               "Keyword Type Values Master": "Master",
               "LC ADD/LESS DETAILS": "Master",
               "MASTER APPROVAL": "Master",
               "Party Master": "Master",
               "Payement Terms Account Mapping": "Master",
               "Payment Term": "Master",
               "PO TEMPLATE": "Master",
               "Power BI Configuration": "Master",
               "Process Operation Master": "Master",
               "Product Sub Category": "Master",
               "Report Designer": "Master",
               "UOM Master": "Master",
               "Vendor Bank Master": "Master",
               "Vendor Ranking": "Master",
               "Vendor Wise Item Rate": "Master",
               "Weighted Average": "Master",
               "Account Log Report": "Reports",
               "Actual Cost Report": "Reports",
               "Adhoc BK Report": "Reports",
               "BOM Vs Indent Vs PO Report": "Reports",
               "Boutique Intl Report": "Reports",
               "Consumed Quantity Report": "Reports",
               "Costing Detail Report": "Reports",
               "General Issue Receive Report": "Reports",
               "IGST Payment Status": "Reports",
               "ITC-04 Goods Received Report": "Reports",
               "ITC-04 Goods Supplied Report": "Reports",
               "Item Lotwise Status": "Reports",
               "Item Lot Wise Tracking Report": "Reports",
               "Lot Wise Report": "Reports",
               "Order Completion Report": "Reports",
               "Outstanding Pending Payment Report": "Reports",
               "Packing List Report": "Reports",
               "PO Followup Report New": "Reports",
               "Production Process Cost Report": "Reports",
               "Production Transaction Raw Consumption Report": "Reports",
               "Stock Transfer Issue/Receive Report": "Reports",
               "Style Stock report": "Reports",
               "Change Password": "Admin",
               "Data Entry Check List": "Admin",
               "Default User Setting": "Admin",
               "Drill Down report": "Admin",
               "Item Wise Rate Check List": "Admin",
               "Login Security Setup": "Admin",
               "Numbering Method": "Admin",
               "Replace Values": "Admin",
               "Replace Values Log": "Admin",
               "Screen Group Master": "Admin",
               "Screen Log": "Admin",
               "ACCESSORY BALANCE FOR ORDER": "Order",
               "Analysis Report": "Order",
               "Bom Approval Itemwise": "Order",
               "Bom Import": "Order",
               "BOM Item Change": "Order",
               "BOM Rate Update": "Order",
               "Bom Revision": "Order",
               "Budget Costing": "Order",
               "Budget Vs Actual": "Order",
               "Costing Approval": "Order",
               "Daily Issue Receive of Fabric and Accessory": "Order",
               "Due Orders Coming Week": "Order",
               "Due Po Details": "Order",
               "Fabric Composition Master": "Order",
               "Final Order Status": "Order",
               "LabDip Complete": "Order",
               "LabDip Creation": "Order",
               "LabDip Dispatch": "Order",
               "LabDip FeedBack": "Order",
               "LabDip Req": "Order",
               "LabDip Verification": "Order",
               "Man Power Cost Analysis": "Order",
               "Merchandising Entry Report": "Order",
               "New Style Sheet Report": "Order",
               "Order Booking (Greige)": "Order",
               "Order Booking Chart": "Order",
               "Order Closing Report ( New )": "Order",
               "Order In Hand": "Order",
               "Order In Hand Report With Image": "Order",
               "Order Status": "Order",
               "Order Status Report": "Order",
               "Order Style Grouping": "Order",
               "Payment Due Report": "Order",
               "Product Catalog": "Order",
               "Product Category Status": "Order",
               "Product Costing": "Order",
               "Purchase Order Ageing Report": "Order",
               "Sale Order (Fabric)": "Order",
               "Sale Orders Log": "Order",
               "SaleOrders Split Detail": "Order",
               "Sales Chart Report": "Order",
               "Shipment Details": "Order",
               "Shipped Orders": "Order",
               "Sub Order Wise Order In Hand": "Order",
               "Summary Order Actual Cost": "Order",
               "Upload Style Image": "Order",
               "Sale Order": "Order",
               "Order Approval": "Order",
               "BOM": "Order",
               "Activity Alert Mapping": "Permission",
               "Approval Configuration": "Permission",
               "Group Permission": "Permission",
               "Permission": "Permission",
               "UCL Limit": "Permission",
               "Users": "Permission",
               "Bank Reference Details": "Commercial",
               "Bill Of Exchange": "Commercial",
               "Cargo Delivery Note": "Commercial",
               "Cdn Query": "Commercial",
               "Company Period Wise Data Freeze": "Commercial",
               "Costing Comparison": "Commercial",
               "Dispatch Advice": "Commercial",
               "Document Printing": "Commercial",
               "DrawBack Detail Report": "Commercial",
               "E-Invoice": "Commercial",
               "Incentive Status Report": "Commercial",
               "Insurance Details": "Commercial",
               "Invoice Detail": "Commercial",
               "Invoice Receiving Details": "Commercial",
               "Invoice Verification": "Commercial",
               "Letter of Credit": "Commercial",
               "Master Register Report": "Commercial",
               "Master Register Report (New)": "Commercial",
               "Month Wise Sale Report": "Commercial",
               "Packing List": "Commercial",
               "Packing List New": "Commercial",
               "Packing List Verification": "Commercial",
               "Pending Sales Bill": "Commercial",
               "Post Shipment Details": "Commercial",
               "Sales Return": "Commercial",
               "Weaving Packing List": "Commercial",
               "Import Activity": "API",
               "Import Actual Production": "API",
               "Import Chart Of Accounts": "API",
               "Import Chart of Cost Center": "API",
               "Import Cost Center Address": "API",
               "Import Design Master": "API",
               "Import HSN Code": "API",
               "Import IPPC Production Planning": "API",
               "Import Item Stock": "API",
               "Import Items": "API",
               "Import Opening Balance Bill Wise": "API",
               "Import Product Category": "API",
               "Import Production Transactions": "API",
               "Import Sales Order": "API",
               "Import the LOB": "API",
               "Retail Receiving Import": "API",
               "Activity Master": "TNA",
               "Calendar": "TNA",
               "FG Status": "TNA",
               "Order Activity": "TNA",
               "Re-assign Responsbility": "TNA",
               "T&A Status": "TNA",
               "T&A Status Report": "TNA",
               "T&A Template": "TNA",
               "TNA Release": "TNA",
               "Un-Complete Activity": "TNA",
               "User Role Masters": "TNA",
               "Buyer Enquiry": "Product Development",
               "Buyer Feedback": "Product Development",
               "Buyer Wise Activity Qty": "Product Development",
               "Material Issue For Product Development": "Product Development",
               "PD Planning Item Wise Approval": "Product Development",
               "Product Development": "Product Development",
               "Product Development Planning": "Product Development",
               "Sample Issue": "Product Development",
               "SAMPLE ISSUE FOR BILL NO": "Product Development",
               "Sample Issue Receive Status Report": "Product Development",
               "Sample Receive": "Product Development",
               "Sample Status": "Product Development",
               "Sample Status Report": "Product Development",
               "Sampling Quotation": "Product Development",
               "ITEM_MASTER": "Report Tool View",
               "MERCHANT": "Report Tool View",
               "SALE ORDER": "Report Tool View"}

    def test_1_Stock_Report(self):
        """
        Performs a series of actions on Swag Labs website to generate and send the Stock Ledger Summary Report.
        """

        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton')

        # Navigate to Stock Ledger Report page
        Sub = self.subMenu['Stock Ledger (New)']
        Men = self.manMenu[Sub]
        self.click_link_text(Men, timeout=30)
        self.click_link_text(Sub, timeout=30)
        self.click_link_text('Stock Ledger (New)', timeout=30)

        # Switch to window
        self.switch_to_window(1)
        # Generate reports for different item groups
        item_groups = {
            '1': ['SVMD247 ACCESSORIES STORE', 'TRIM'],
            '2': ['SVMD247 ACCESSORIES STORE', 'PACKING'],
            '3': ['SVMD247 FABRIC STORE', 'FABRIC']
        }

        for key, value in item_groups.items():
            location = value[0]  # First element in the list is the location
            item_group = value[1]
            # Select options
            self.select_option_by_text('[name="locationID"]', location, timeout=30)
            self.select_option_by_text('[name="ledgerTypeID"]', 'Store Wise Summary', timeout=30)
            self.select_option_by_text('[name="item_group_id"]', item_group, timeout=30)

            # Select start date
            self.click('[name="StartDate"]', timeout=30)
            self.click('.btn-clear-wrapper', timeout=30)
            self.click('[name="StartDate"]', timeout=30)
            self.select_target_date_Jquery(
                'div.bs-datepicker-body > table > tbody > tr > td > span[class="ng-star-inserted"]')

            # Select end date
            self.click('input[name="EndDate"]', timeout=30)
            self.click('.btn-clear-wrapper', timeout=30)
            self.click('input[name="EndDate"]', timeout=30)
            self.select_target_date_Jquery(
                'div.bs-datepicker-body > table > tbody > tr > td > span[class="ng-star-inserted"]')

            # Click on generate button
            self.click('[class="btn btn-info btn-sm"]', timeout=30)
            self.wait_for_element_present('.row.ng-star-inserted .ag-root-wrapper-body .ag-row.ag-row-odd', timeout=30)
            self.wait_for_element_not_visible('body > div.block-ui-container.ng-scope', timeout=30)

            # Click on export button
            self.click('[class="btn btn-primary btn-sm"]', timeout=30)
            self.wait_for_element_not_visible('body > div.block-ui-container.ng-scope', timeout=30)

            # Wait for download
            wait_time_seconds = 8
            time.sleep(wait_time_seconds)

            # Rename and format the downloaded file
            File_Path = 'downloaded_files'
            FormatingStockLedger(File_Path + '\\Summary.xlsx', File_Path + f'\\{item_group.title()} Stock Report.xlsx')

            self.refresh_page()

        File_Path = 'downloaded_files'

        # Send email with reports attached
        today_Date = datetime.now().strftime('%#d-%b-%Y')
        recipients = ('erp.svm@svmincorporation.com', 'naman@svmincorporation.com', 'sanjeev@svmincorporation.com',
                      'erpsupport@svmincorporation.com', 'store.svm@gmail.com', 'svmfabric@yahoo.com',
                      'svm.printing@svmincorporation.com', 'store.svm@gmail.com',)
        # recipients = ['erpsupport@svmincorporation.com']
        subject = 'Stock Report ' + today_Date
        Body = '''<html>
            <head></head>
            <body>
                <p style="margin: 0;">Dear All</p>
                <br>
                <p style="margin: 0;">Please find attached the Stock Report for your review and comments:</p>
                <br>
                <p style="margin: 0;">Thanks & Regards</p>
                <p style="margin: 0;">Shashi Giri (ERP)</p>
            </body>
            </html>'''
        filesName = ['Packing Stock Report.xlsx', 'Trim Stock Report.xlsx', 'Fabric Stock Report.xlsx']
        Send_selenium_report(recipients, Body, File_Path, subject, filesName)

        # Remove all generated report files
        os.remove(File_Path + '\\Trim Stock Report.xlsx')
        os.remove(File_Path + '\\Packing Stock Report.xlsx')
        # os.remove(File_Path + '\\Fabric Stock Report.xlsx')

    def test_2_Production_Detail_Report(self):
        """
        Performs a series of actions on Swag Labs website to generate and send the Production Detail Report.
        """

        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton', timeout=30)

        # Navigate to 'Production Detail Report' screen
        ScreenName = 'Production Detail Report'
        Sub = self.subMenu[ScreenName]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(ScreenName, timeout=30)

        wait_time_seconds = 1
        time.sleep(wait_time_seconds)

        # Select location
        self.click('div#location > a', timeout=30)
        self.click('div#location li:nth-child(6)', timeout=30)

        # Select start date
        self.click('[admincontrol="main_model.from_date"]>[class="btn btn-primary ng-scope"]', timeout=30)
        self.click('[class="btn btn-sm btn-danger ng-binding"]', timeout=30)
        wait_time_seconds = 1
        time.sleep(wait_time_seconds)
        self.click('[admincontrol="main_model.from_date"]>[class="btn btn-primary ng-scope"]', timeout=30)
        self.select_target_date_Basic(
            'div[ng-switch="datepickerMode"] > table > tbody > tr > td > button > span[class="ng-binding"]')

        # Select end date
        self.click('[admincontrol="main_model.to_date"]>[class="btn btn-primary ng-scope"]', timeout=30)
        self.click('[class="btn btn-sm btn-danger ng-binding"]', timeout=30)
        wait_time_seconds = 1
        time.sleep(wait_time_seconds)
        self.click('[admincontrol="main_model.to_date"]>[class="btn btn-primary ng-scope"]', timeout=30)
        self.select_target_date_Basic(
            'div[ng-switch="datepickerMode"] > table > tbody > tr > td > button > span[class="ng-binding"]')

        # Click on print and export to PDF buttons
        self.click('[id="print1"]', timeout=30)
        wait_time_seconds = 30
        time.sleep(wait_time_seconds)

        # Rename and format the downloaded file
        FileName = 'Production Detail Report'
        File_Path = 'downloaded_files'
        rename_latest_file(File_Path, FileName)

        # Click on WIP Print and export to Excel buttons
        self.click('[id="wip_export"]', timeout=30)
        wait_time_seconds = 35
        time.sleep(wait_time_seconds)

        # Rename and format the downloaded file
        FileName = 'Production Detail Report'
        rename_latest_file(File_Path, FileName)
        Change_File_Ext(File_Path)
        Production_Detail_Report_Formating(File_Path + '\\Production Detail Report.xlsx')

        # Get today's date
        today_Date = datetime.now().strftime('%#d-%b-%Y')

        # Define recipients, subject, and body for email
        recipients = ('erp.svm@svmincorporation.com', 'naman@svmincorporation.com',
                      'sanjeev@svmincorporation.com', 'erpsupport@svmincorporation.com',
                      'PRODUCTION@SVMINCORPORATION.COM', 'svmproduction247@gmail.com',
                      'prince@svmincorporation.com', 'merch.svm8@svmincorporation.com',
                      'MERCH.SVM3@SVMINCORPORATION.COM', 'GARMENTS@SVMINCORPORATION.COM',
                      'merchant@svmincorporation.com', 'merch.svm10@svmincorporation.com')
        # recipients = ['erpsupport@svmincorporation.com']
        subject = 'Production Detail Report ' + today_Date
        Body = '''<html>
            <head></head>
            <body>
                <p style="margin: 0;">Dear All</p>
                <br>
                <p style="margin: 0;">Please find attached the Production Detail Report for your review and comments:</p>
                <br>
                <p style="margin: 0;">Thanks & Regards</p>
                <p style="margin: 0;">Shashi Giri (ERP)</p>
            </body>
            </html>'''

        # Send email with report attached
        FileName = ['Production Detail Report.pdf', 'Production Detail Report.xlsx']
        Send_selenium_report(recipients, Body, File_Path, subject, FileName)

        # Remove the generated report files
        os.remove(File_Path + '\\Production Detail Report.xlsx')
        os.remove(File_Path + '\\Production Detail Report.pdf')
        os.remove(File_Path + '\\Fabric Stock Report.xlsx')

    def test_3_Procurement_Status(self):
        """
        Performs a series of actions on Swag Labs website to generate and send the Procurement Status Report.
        """

        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton', timeout=30)

        # Navigate to 'Procurement Status New' screen
        ScreenName = 'Procurement Status New'
        Sub = self.subMenu[ScreenName]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(ScreenName, timeout=30)

        # Switch to window
        self.switch_to_window(1, timeout=30)

        # Select options for Procurement Status Report
        self.click('[aria-owns="bs-select-4"]', timeout=30)
        self.click_link_text("Fabric", timeout=30)
        self.click("app-procurement-status .form-row:nth-of-type(3) "
                   "[class='form-group col-md-2']:nth-of-type(1) .input-with-post-icon", timeout=30)
        self.click(".btn.btn-success.ng-tns-c28-1", timeout=30)
        self.click("app-procurement-status .form-row:nth-of-type(4) "
                   "[class='form-group col-md-2']:nth-of-type(1) .input-with-post-icon", timeout=30)
        self.click(".btn.btn-success.ng-tns-c28-2", timeout=30)
        self.click("app-procurement-status .form-row:nth-of-type(4) "
                   "[class='form-group col-md-2']:nth-of-type(2) .input-with-post-icon", timeout=30)
        self.click(".btn.btn-success.ng-tns-c28-3", timeout=30)
        self.click("app-procurement-status .form-row:nth-of-type(2) "
                   "[class='form-group col-md-2']:nth-of-type(2) > div", timeout=30)
        self.click("[x-placement] li:nth-of-type(2) [tabindex]", timeout=30)
        self.click("[class] .btn-sm:nth-of-type(2)", timeout=30)

        # Wait for elements to be present
        self.wait_for_element_present("[ref='eBodyViewport'] [ref='eCenterColsClipper'] [role='row']", timeout=30)

        # Click on buttons
        self.click("app-procurement-status .card:nth-child(3) > div:nth-of-type(2) .form-control-sm", timeout=30)
        self.click('[class="btn btn-success btn-sm"]', timeout=30)

        # Wait for the report to be generated
        wait_time_seconds = 8
        time.sleep(wait_time_seconds)

        # Rename and format the downloaded file
        FileName = 'Procurement Status Report'
        File_Path = 'downloaded_files'
        rename_latest_file(File_Path, FileName)
        self.driver.close()
        self.switch_to_default_window()
        Procurement_Status_Report_Formating(File_Path + '\\Procurement Status Report.xlsx')

        # Get today's date
        today_Date = datetime.now().strftime('%#d-%b-%Y')

        # Define recipients, subject, and body for email
        recipients = ('erp.svm@svmincorporation.com', 'naman@svmincorporation.com', 'sanjeev@svmincorporation.com',
                      'erpsupport@svmincorporation.com', 'store.svm@gmail.com', 'svmfabric@yahoo.com',
                      'svm.printing@svmincorporation.com', 'store.svm@gmail.com')

        # recipients = ['erpsupport@svmincorporation.com']
        subject = 'Procurement Status Report ' + today_Date
        Body = '''<html>
            <head></head>
            <body>
                <p style="margin: 0;">Dear All</p>
                <br>
                <p style="margin: 0;">Please find attached the Procurement Status for your review and comments:</p>
                <br>
                <p style="margin: 0;">Thanks & Regards</p>
                <p style="margin: 0;">Shashi Giri (ERP)</p>
            </body>
            </html>'''
        filesName = ['Procurement Status Report.xlsx']

        # Send email with report attached
        Send_selenium_report(recipients, Body, File_Path, subject, filesName)

        # Remove the generated report file
        os.remove(File_Path + '\\Procurement Status Report.xlsx')

    def test_4_PO_Followup_Report(self):
        """
        Performs a series of actions on Swag Labs website to generate and send the PO Followup Report.
        """

        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton', timeout=30)

        # Navigate to 'PO Followup Report' screen
        ScreenName = 'PO Followup Report'
        Sub = self.subMenu[ScreenName]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(ScreenName, timeout=30)

        # Select options for PO Followup Report
        self.click('[id="is_show"]', timeout=30)
        self.click('[id="item_group_type"]', timeout=30)
        self.click_link_text("FABRIC", timeout=30)
        self.click('[id="po_type"]', timeout=30)
        self.click('#po_type > div > ul > li:nth-child(3)', timeout=30)
        self.click('[id="is_bal_per"]', timeout=30)
        self.clear('[id="bal_per"]', timeout=30)
        self.send_keys('[id="bal_per"]', 10, timeout=30)

        # Click on search and wait for elements to be present
        self.click('#SearchList', timeout=30)

        # Export the report to Excel
        self.click('#ExportToExcel', timeout=30)
        wait_time_seconds = 15
        time.sleep(wait_time_seconds)

        # Rename and format the downloaded file
        FileName = 'PO Followup'
        File_Path = 'downloaded_files'
        rename_latest_file(File_Path, FileName)
        Change_File_Ext(File_Path)
        PO_Followup_Report_Formating(File_Path + '\\PO Followup.xlsx', File_Path + '\\PO Followup Report.xlsx')

        # Get today's date
        today_Date = datetime.now(timezone.utc).strftime('%#d-%b-%Y')

        # Define recipients, subject, and body for email
        recipients = ('erp.svm@svmincorporation.com', 'naman@svmincorporation.com', 'sanjeev@svmincorporation.com',
                      'erpsupport@svmincorporation.com', 'store.svm@gmail.com', 'svmfabric@yahoo.com',
                      'svm.printing@svmincorporation.com', 'store.svm@gmail.com', 'prince@svmincorporation.com')
        subject = 'PO Followup Report ' + today_Date
        Body = '''<html>
            <head></head>
            <body>
                <p style="margin: 0;">Dear All</p>
                <br>
                <p style="margin: 0;">Please find attached the PO Followup Report for your review and comments:</p>
                <br>
                <p style="margin: 0;">Thanks & Regards</p>
                <p style="margin: 0;">Shashi Giri (ERP)</p>
            </body>
            </html>'''
        filesName = ['PO Followup Report.xlsx']

        # Send email with report attached
        Send_selenium_report(recipients, Body, File_Path, subject, filesName)

        # Remove the generated report file
        os.remove(File_Path + '\\PO Followup Report.xlsx')

    def test_5_Order_Completion_Report(self):
        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton', timeout=30)

        # Navigate to 'Order Completion Report' screen
        ScreenName = 'Order Completion Report'
        Sub = self.subMenu[ScreenName]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(ScreenName, timeout=30)

        # Select 'SVM INC' company and 'unship' status
        self.click('#company', timeout=30)
        self.click_link_text('SVM INC', timeout=30)
        self.click('#status', timeout=30)
        self.click_link_text('unship', timeout=30)

        # Click on search and wait for elements to be present
        self.click('#SearchList', timeout=30)
        self.wait_for_element_present('.ui-grid-render-container-body .ui-grid-canvas > div', timeout=30)

        # Select all elements and click on export
        self.click('div[ng-if="grid.options.enableSelectAll"]', timeout=30)
        self.click('i.ui-grid-icon-menu', timeout=30)
        self.click('li#menuitem-1>button', timeout=30)

        # Wait for file download and rename file
        wait_time_seconds = 10
        time.sleep(wait_time_seconds)
        File_Path = "downloaded_files"
        FileName = "Completion"
        rename_latest_file(File_Path, FileName)
        Change_File_Ext(File_Path)

        # Format and rename Order Completion Report
        start_date = datetime.now(timezone.utc).date()
        Order_Completion_Report_Formating(File_Path + '\\Completion.xlsx', File_Path + '\\Order Completion Report.xlsx')

        # Send email with report attached
        recipients = ['erpsupport@svmincorporation.com', 'naman@svmincorporation.com', 'dabasakash44@gmail.com',
                      'svmproduction247@gmail.com', 'store.svm@gmail.com']
        subject = 'Up Coming and Order Completion Report ' + str(start_date.strftime('%#d-%b-%Y'))
        Body = '''<html>
            <head></head>
            <body>
                <p style="margin: 0;">Dear All</p>
                <br>
                <p style="margin: 0;">Please find attached the Up Coming and Order Completion Report for your review and comments:</p>
                <br>
                <p style="margin: 0;">Thanks & Regards</p>
                <p style="margin: 0;">Shashi Giri (ERP)</p>
            </body>
            </html>'''
        filesName = ['Order Completion Report.xlsx']
        Send_selenium_report(recipients, Body, File_Path, subject, filesName)

        # Remove the generated report file
        os.remove(File_Path + '\\Order Completion Report.xlsx')

    def test_6_SubOrder_WiseInHand_Report(self):
        # Login to Swag Labs
        self.login_to_swag_labs()

        # Click on top button
        self.click('#topButton', timeout=30)

        # Navigate to 'Sub Order Wise Order In Hand' screen
        Screen = 'Sub Order Wise Order In Hand'
        Sub = self.subMenu[Screen]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(Screen, timeout=30)

        # Select 'SVM INC' company
        self.click('#company', timeout=30)
        self.click_link_text('SVM INC', timeout=30)

        # Click on 'order_summary'
        self.click('#order_summary', timeout=30)
        self.click('#Search', timeout=30)

        # Wait for element to be present
        Path = ("#gridOrderInHand > div > div> div > div > div.dx-scrollable-content > div > table > tbody > "
                "tr.dx-row.dx-data-row.dx-row-lines.dx-column-lines > td")
        self.wait_for_element_present(Path, timeout=50)

        # Click on 'Excel' button to export data
        self.click('#Excel', timeout=30)

        # Wait for file download and renaming
        wait_time_seconds = 5
        time.sleep(wait_time_seconds)
        File_Path = "downloaded_files"
        FileName = "sourceWB"
        rename_latest_file(File_Path, FileName)
        Change_File_Ext(File_Path)

        # Click on top button again
        self.click('#topButton', timeout=30)

        # Navigate to 'MERCHANT' screen
        Screen = 'MERCHANT'
        Sub = self.subMenu[Screen]
        Men = self.manMenu[Sub]
        self.click_link_text(Men)
        self.click_link_text(Sub)
        self.click_link_text(Screen, timeout=30)

        # Click on 'search' button
        self.click('#search', timeout=30)

        # Wait for element to be present
        Path = ("#gridContainer_Detail > div > div.dx-datagrid-rowsview.dx-datagrid-nowrap.dx-scrollable.dx-scrollable-"
                "customizable-scrollbars.dx-scrollable-both.dx-scrollable-simulated.dx-visibility-change-handler > div "
                "> div > div.dx-scrollable-content > div > table > tbody > tr")
        self.wait_for_element_present(Path, timeout=50)

        # Click on export button
        self.click("[class='dx-icon dx-icon-export-to']", timeout=30)
        self.click("[data-bind] .dx-menu-item-wrapper:nth-of-type(1) .dx-menu-item-text", timeout=30)

        # Wait for file download
        wait_time_seconds = 5
        time.sleep(wait_time_seconds)

        # Call the function with the desired file path
        FormatingOrderInHand(File_Path)

        start_date = datetime.now(timezone.utc).date()

        recipients = ('erp.svm@svmincorporation.com', 'naman@svmincorporation.com', 'sanjeev@svmincorporation.com',
                      'erpsupport@svmincorporation.com', 'prince@svmincorporation.com', 'cfo@svmincorporation.com')

        # recipients = ['erpsupport@svmincorporation.com']
        subject = 'Order InHand Report ' + str(start_date.strftime('%#d-%b-%Y'))
        Body = '''<html>
                <head></head>
                <body>
                    <p style="margin: 0;">Dear All</p>
                    <br>
                    <p style="margin: 0;">Please find attached the Order InHand Report for your review and comments:</p>
                    <br>
                    <p style="margin: 0;">Thanks & Regards</p>
                    <p style="margin: 0;">Shashi Giri (ERP)</p>
                </body>
                </html>'''
        filesName = ['Order InHand Report.xlsx']
        Send_selenium_report(recipients, Body, File_Path, subject, filesName)

        # Remove the generated report file
        os.remove(File_Path + '\\Order InHand Report.xlsx')
        # os.remove(File_Path + '\\ReportData')
        # os.remove(File_Path + '\\sourceWB')
